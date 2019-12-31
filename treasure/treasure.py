#!/user/bin/env python3
# -*- coding: utf-8 -*-

from openpyxl.reader.excel import load_workbook
from openpyxl.styles import Font, colors
import time
import sys


class Product:

    def __init__(self, destination_name, product_id, poi_id):
        self.destination_name = destination_name
        self.product_id = product_id
        self.poi_id = poi_id


class Route:

    def __init__(self, product_city, poi_id):
        self.product_city = product_city
        self.poi_id = poi_id


treasure = {}
poi = {}
treasure_match = {}
sheet = None
wb = None


def load_treasure(filename):
    """
    加载藏宝图
    :param filename:
    :return:
    """
    begin = time.time()
    global wb
    wb = load_workbook(filename)
    end = time.time()
    print(" read time {}", end - begin)

    sheets = wb.sheetnames
    global sheet
    sheet = wb[sheets[0]]

    for index, row in enumerate(sheet.__iter__()):
        if index == 0:
            continue
        key = str(row[6].value) + "+" + str(row[4].value)
        route_data = Route(row[4].value, row[6].value)
        treasure[key] = route_data


def load_poi(filename):
    """
    加载poi
    :param filename:
    :return:
    """
    wb_poi = load_workbook(filename, read_only=True)
    sheets = wb_poi.sheetnames
    sheet_poi = wb_poi[sheets[0]]
    for index, row in enumerate(sheet_poi.__iter__()):
        if index == 0:
            continue
        key = row[7].value
        product_data = Product(row[3].value, row[7].value, row[9].value)
        poi[key] = product_data


def match():
    for key in treasure:
        key_list = str(key).split("+")

        for poi_key in poi:
            product_data = poi[poi_key]

            distinct = str(product_data.destination_name) + "+" + str(product_data.poi_id)
            match_flag = False

            for treasure_key in key_list:
                if str(treasure_key) in str(distinct):
                    match_flag = True
                else:
                    match_flag = False
                    break
            if match_flag:
                try:
                    value = treasure_match[key]
                    value.append(product_data.product_id)
                    treasure_match[key] = value
                except KeyError:
                    treasure_match[key] = [product_data.product_id]


def write_treasure(filename):
    """
    把 treasure_match 写入 treasure
    :param filename:
    :return:
    """

    content_font = Font(name="微软雅黑", size=10, color=colors.RED)
    for index, row in enumerate(sheet.__iter__()):
        if index == 0:
            continue
        position_1 = 'H' + str(index + 1)
        position_2 = 'I' + str(index + 1)
        position_3 = 'J' + str(index + 1)
        content_position_1 = 'G' + str(index + 1)
        content_position_2 = 'E' + str(index + 1)
        match_id = str(sheet[content_position_1].value) + "+" + str(sheet[content_position_2].value)

        try:
            content = treasure_match[match_id]
            sheet[position_1].font = content_font
            sheet[position_1] = 'T'

            sheet[position_2].font = content_font
            sheet[position_2] = len(content)

            sheet[position_3].font = content_font
            sheet[position_3] = str(content)

        except KeyError:
            if sheet['A' + str(index + 1)].value is None:
                continue
            sheet[position_1].font = content_font
            sheet[position_1] = 'F'

    begin = time.time()
    wb.save(filename)
    end = time.time()
    print(" write time {}", end - begin)


if __name__ == '__main__':
    if len(sys.argv) != 3:
        print("enter treasure and product")
    else:
        origin = sys.argv[1]
        target = sys.argv[2]
        load_treasure("./" + str(origin) + ".xlsx")
        load_poi("./" + str(target) + ".xlsx")
        match()
        write_treasure("./" + str(origin) + ".xlsx")
        treasure.clear()
        poi.clear()
        treasure_match.clear()
