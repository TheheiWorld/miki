#!/user/bin/env python3
# -*- coding: utf-8 -*-

from openpyxl.reader.excel import load_workbook
from openpyxl.styles import Font, colors
import time
import sys


class Product:

    def __init__(self, name, poi_id):
        self.name = name
        self.poi_id = poi_id


poi = {}

def load_poi(filename):
    """
    加载poi
    :param filename:
    :return:
    """
    wb_poi = load_workbook(filename, read_only=True)
    sheets = wb_poi.sheetnames
    sheet_poi = wb_poi[sheets[0]]
    for index, row in enumerate(sheet_poi.__iter__()):
        if index == 0:
            continue
        key = str(row[2].value)
        route_data = Product(row[2].value, row[1].value)
        poi[key] = route_data


def write_treasure(filename):
    wb_poi = load_workbook(filename)
    sheets = wb_poi.sheetnames
    sheet = wb_poi[sheets[0]]

    content_font = Font(name="微软雅黑", size=10, color=colors.RED)
    for index, row in enumerate(sheet.__iter__()):
        if index < 2:
            continue
        name = str(row[9].value)
        name_list = name.split("+")

        position_1 = 'N' + str(index + 1)
        poi_id_list = ''
        for n in name_list:
            print(n)
            product = poi.get(n)
            if product is None:
                poi_id_list = poi_id_list + "+ Null"
            else:
                poi_id_list = poi_id_list + "+ " + str(product.poi_id)
        poi_id_list = poi_id_list[1::]
        sheet[position_1].font = content_font
        sheet[position_1] = str(poi_id_list)
    wb_poi.save(filename)


if __name__ == '__main__':
    load_poi("/Users/juststand/Desktop/poi.xlsx")
    write_treasure("/Users/juststand/Desktop/detail.xlsx")
